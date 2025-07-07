import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO

def load_reactor_data(filepath):
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip().str.lower()
    rename_map = {
        "vessel id": "reactor id",
        "min sensing volume": "min sensing",
        "min stirring volume": "min stirring",
        "capacity": "max volume",
        "moc": "moc", 
        "utilities": "utilities",
        "agitator": "agitator"
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    if "utilities" not in df.columns:
        st.error(" 'utilities' column not found in the uploaded Excel. Please ensure it's named correctly.")
        return pd.DataFrame()

    df["moc"] = df["moc"].str.upper().replace({"ALL GLASS": "GLR"})
    df["materials"] = df["moc"].apply(lambda x: [m.strip() for m in x.split("/")])
    df["thermal options"] = df["utilities"].astype(str).apply(lambda x: [t.strip().upper() for t in x.split(",")])
    df["agitator"] = df["agitator"].astype(str).str.upper()
    return df[["reactor id", "min sensing", "min stirring", "max volume", "materials", "thermal options", "agitator"]]

def load_filter_data(uploaded_file):
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip().str.lower()
    return df

def load_dryer_data(uploaded_file):
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip().str.lower()  # Clean column names
    rename_map = {
        "dryer id": "equipment id",   # Standardized for consistency
        "capacity": "capacity",
        "moc": "moc",
        "dryer type": "dryer type"
    }
    df = df.rename(columns=rename_map)
    return df

def collect_unit_operation(unit_op_id):
    steps = []
    total_volume = 0
    first_step_volume = None

    st.subheader("Add Steps to Unit Operation")
    step_count = 0
    add_more_steps = True

    while add_more_steps:
        step_count += 1
        st.markdown(f"### Step {step_count}")
        operation = st.selectbox(f"Select operation type for Step {step_count}", ["charge", "addition"], key=f"op_{unit_op_id}_{step_count}")
        material = st.selectbox(f"Select material type for Step {step_count}", ["reagent 1", "reagent 2", "reagent 3", "KSM", "solvent"], key=f"mat_{unit_op_id}_{step_count}")
        volume = st.number_input(f"Enter volume (L) for Step {step_count}", min_value=0.0, key=f"vol_{unit_op_id}_{step_count}")

        actual_volume = volume
        if material == "KSM":
            percentage = st.number_input(f"Enter percentage of KSM for Step {step_count}", min_value=0.0, max_value=100.0, key=f"KSM_{unit_op_id}_{step_count}")
            if percentage > 0:
                actual_volume = volume / (percentage / 100)

        if first_step_volume is None:
            first_step_volume = actual_volume

        total_volume += actual_volume
        steps.append({
            "unit_op": unit_op_id,
            "step": step_count,
            "operation": operation,
            "material": material,
            "input_volume": volume,
            "actual_volume": actual_volume,
            "accumulated_volume": total_volume
        })

        add_more_steps = st.radio(f"Add another step after Step {step_count}?", ["yes", "no"], index=1, key=f"cont_{unit_op_id}_{step_count}") == "yes"

    return first_step_volume, total_volume, steps

def filter_reactors(df, user_input, first_step_vol, total_vol):
    df = df[(df["min sensing"] <= first_step_vol) & (df["min stirring"] <= first_step_vol)]

    process_type = user_input["process_type"]
    vol_limit = 0.7 if process_type in ["distillation", "reaction", "pressurized"] else 0.95
    df = df[df["max volume"] * vol_limit >= total_vol]

    if user_input["ph_condition"] == "basic":
        allowed = ["SSR", "HAR"]
    elif user_input["ph_condition"] == "acidic":
        allowed = ["GLR", "HAR"]
    elif user_input["ph_condition"] == "neutral":
        allowed = ["GLR", "SSR", "HAR"]
    elif user_input["ph_condition"] == "coupon":
        mat = user_input["coupon_materials"][0].strip().upper()
        if user_input["corrosion_rate"] < 0.1:
            allowed = [mat]
        else:
            st.error("Corrosion rate too high for this material.")
            return pd.DataFrame()
    else:
        allowed = []

    df = df[df["materials"].apply(lambda mats: any(m in mats for m in allowed))]

    temp = user_input["temperature"]
    if 10 <= temp <= 20:
        thermal = ["CHB"]
    elif 20 < temp <= 35:
        thermal = ["CT"]
    elif 20 < temp <= 90:
        thermal = ["HW"]
    else:
        thermal = ["LPS", "HOT OIL", "EJECTION CONDENSATE"]

    df = df[df["thermal options"].apply(lambda opts: any(t in opts for t in thermal))]

    preferred = []
    if user_input["reaction_nature"] == "homogeneous":
        preferred = ["PROPELLOR", "PBT", "RCI", "ANCHOR", "CBRT"]
    elif user_input["reaction_nature"] == "heterogeneous":
        if user_input["reaction_subtype"] == "biphasic":
            preferred = ["PROPELLOR", "PBT", "CBRT", "RCI"]
        elif user_input["reaction_subtype"] == "solid-liquid":
            preferred = ["PROPELLOR", "PBT", "CBRT", "RCI", "ANCHOR"]
        elif user_input["reaction_subtype"] == "gas-liquid":
            preferred = ["RUSTON", "DISC"]

    df["Preference Match"] = df["agitator"].apply(lambda a: "yes" if any(p in a for p in preferred) else "warning")
    return df

def filter_filters(df, user_input, filter_types_required):
    ph_condition = user_input["ph_condition"]

    if ph_condition == "basic":
        allowed = ["SSR", "HAR", "HALAR"]
    elif ph_condition == "acidic":
        allowed = ["HALAR", "HAR"]
    elif ph_condition == "neutral":
        allowed = ["SSR", "HAR", "HALAR"]
    elif ph_condition == "coupon":
        mat = user_input["coupon_materials"][0].strip().upper()
        if user_input["corrosion_rate"] < 0.1:
            allowed = [mat]
        else:
            st.error("Corrosion rate too high for this material.")
            return pd.DataFrame()
    else:
        allowed = []

    df = df[df["moc"].astype(str).str.upper().isin(allowed)]

    # Volume calculation
    mass = user_input["mass"]
    bulk_density = user_input["bulk_density"]
    volume_m3 = mass / bulk_density if bulk_density > 0 else 0
    volume_litres = volume_m3 * 1000
    st.write(f"Volume required (L): {volume_litres:.2f}")

    if "cake capacity" not in df.columns:
        st.error("'cake capacity' column not found in the uploaded Excel.")
        return pd.DataFrame()

    df = df[df["cake capacity"] * 0.9 >= volume_litres]

    if not filter_types_required:
        st.warning("No filter type matched the selected filter property.")
        return pd.DataFrame()

    if "filter type" not in df.columns:
        st.error("'filter type' column not found in the uploaded Excel.")
        return pd.DataFrame()

    df["filter type"] = df["filter type"].astype(str).str.upper()
    df = df[df["filter type"].apply(lambda x: any(f in x for f in filter_types_required))]

    return df


def filter_dryers(df, user_input):
    ph_condition = user_input["ph_condition"]

    # MOC compatibility based on pH
    if ph_condition == "basic":
        allowed = ["SSR", "HAR", "HALAR"]
    elif ph_condition == "acidic":
        allowed = ["HALAR", "HAR"]
    elif ph_condition == "neutral":
        allowed = ["SSR", "HAR", "HALAR"]
    elif ph_condition == "coupon":
        mat = user_input["coupon_materials"][0].strip().upper()
        if user_input["corrosion_rate"] < 0.1:
            allowed = [mat]
        else:
            st.error("Corrosion rate too high for this material.")
            return pd.DataFrame()
    else:
        allowed = []

    # Filter by MOC
    df = df[df["moc"].astype(str).str.upper().isin(allowed)]

    # Volume check
    volume_L = user_input["volume"]
    st.write(f"Volume required (L): {volume_L:.2f}")

    if "capacity" not in df.columns:
        st.error("'capacity' column not found in the uploaded Excel.")
        return pd.DataFrame()

    # Capacity should be >= required volume (with 90% margin)
    df = df[df["capacity"] * 0.9 >= volume_L]

    return df




def export_steps_to_excel(steps_by_unitop):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        all_steps = []
        for unitop_id, (steps, selected_reactor) in enumerate(steps_by_unitop, start=1):
            for s in steps:
                all_steps.append({
                    "Unit Operation": unitop_id,
                    "Operation": s["operation"],
                    "Material": s["material"],
                    "Volume Added (L)": s["actual_volume"],
                    "Accumulated Volume (L)": s["accumulated_volume"],
                    "Reactor ID": selected_reactor
                })

        df_export = pd.DataFrame(all_steps)
        df_export.to_excel(writer, index=False, sheet_name="Steps", startrow=2)
        ws = writer.sheets["Steps"]

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        current_row = 3
        prev_op = None
        for i, row in enumerate(df_export.itertuples(index=False), start=current_row):
            if prev_op != row[0]:
                start_row = i
                count = df_export["Unit Operation"].tolist().count(row[0])
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + count - 1, end_column=1)
                cell = ws.cell(row=start_row, column=1)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                prev_op = row[0]

    buffer.seek(0)
    return buffer

def main():
    st.set_page_config("Reactor and Filter Selector", layout="wide")
    st.title("Process Engineering Automation")

    # Sidebar for tracking selections
    if "selections" not in st.session_state:
        st.session_state.selections = []

    with st.sidebar:
        st.header("Unit Operation Steps")
        if st.session_state.selections:
            for i, (step_log, selection) in enumerate(st.session_state.selections):
                st.markdown(f"### Step {i+1}")
                for step in step_log:
                    st.markdown(f"- **Operation:** {step.get('operation', 'N/A')}")
                    st.markdown(f"- **Material:** {step.get('material', 'N/A')}")
                    st.markdown(f"- **Input Volume (L):** {step.get('input_volume', 0)}")
                    st.markdown(f"- **Actual Volume (L):** {step.get('actual_volume', 0)}")
                    st.markdown(f"- **Accumulated Volume (L):** {step.get('accumulated_volume', 0)}")
                st.markdown(f"**Selected Equipment:** {selection}")
                if st.button(f"❌ Remove Step {i+1}", key=f"remove_{i}"):
                    st.session_state.selections.pop(i)
                    st.rerun()
        else:
            st.info("No unit operations added yet.")

    uploaded_file = st.file_uploader("Upload reactor database", type="xlsx")
    if not uploaded_file:
        st.info("Upload the reactor database to start.")
    else:
        df = load_reactor_data(uploaded_file)
        if not df.empty:
            for idx in range(len(st.session_state.selections), len(st.session_state.selections) + 1):
                st.header("Enter Process Conditions")
                batch_id = idx + 1
                st.markdown(f"## Unit Operation {batch_id}")

                unit_op_type = st.selectbox("Select unit operation type", ["reaction", "distillation", "pressurized", "extraction/workup", "filtration", "drying"], key=f"unit_type_{batch_id}")
                ph_condition = st.selectbox("pH condition", ["basic", "acidic", "neutral", "coupon"], key=f"ph_{batch_id}")
                corrosion_rate = 0
                coupon_materials = []
                if ph_condition == "coupon":
                    corrosion_rate = st.number_input("Corrosion rate (mm/year)", min_value=0.0, key=f"cr_{batch_id}")
                    coupon_materials = [st.text_input("Material for coupon study", key=f"mat_{batch_id}").upper()]

                temperature = st.number_input("Process temperature (°C)", min_value=0.0, key=f"temp_{batch_id}")

                # ---------- NON-FILTRATION AND NON-DRYING OPERATIONS ----------
                if unit_op_type not in ["filtration", "drying"]:
                    reaction_nature = st.selectbox("Nature of reaction", ["none", "homogeneous", "heterogeneous"], key=f"rn_{batch_id}")
                    reaction_subtype = None
                    if reaction_nature == "heterogeneous":
                        reaction_subtype = st.selectbox("Subtype", ["biphasic", "solid-liquid", "gas-liquid"], key=f"rs_{batch_id}")
                    st.markdown("---")
                    first_vol, total_vol, step_log = collect_unit_operation(batch_id)
                    if st.button(f"Submit Unit Operation {batch_id}", key=f"submit_{batch_id}"):
                        user_input = {
                            "process_type": unit_op_type,
                            "ph_condition": ph_condition,
                            "corrosion_rate": corrosion_rate,
                            "coupon_materials": coupon_materials,
                            "temperature": temperature,
                            "reaction_nature": reaction_nature,
                            "reaction_subtype": reaction_subtype
                        }
                        matched_df = filter_reactors(df.copy(), user_input, first_vol, total_vol)
                        if not matched_df.empty:
                            styled = matched_df[["reactor id", "min sensing", "min stirring", "max volume", "agitator", "Preference Match"]]
                            st.success(f"Reactors matching Unit Operation {batch_id}")
                            selected_reactor = st.selectbox("Select one reactor to use:", styled["reactor id"].tolist(), key=f"sel_reactor_{batch_id}")
                            st.dataframe(styled.style.applymap(
                                lambda v: "background-color: #d4edda" if v == "yes" else "background-color: #fff3cd",
                                subset=["Preference Match"]
                            ))
                            st.session_state.selections.append((step_log, selected_reactor))
                        else:
                            st.warning("No matching reactors found for this unit operation.")

                # ---------- FILTRATION OPERATION ----------
                elif unit_op_type == "filtration":
                    uploaded_filter_file = st.file_uploader(f"Upload Filter Database (for Unit Operation {batch_id})", type=["xlsx"], key=f"upload_filter_{batch_id}")

                    st.markdown("### Or manually enter custom filter (not in database)")
                    custom_filter = st.checkbox("Enter custom filter details", key=f"custom_filter_chk_{batch_id}")
                    if custom_filter:
                        custom_filter_name = st.text_input("Filter ID or Name", key=f"custom_filter_name_{batch_id}")
                        custom_volume = st.number_input("Filtered Volume (L)", min_value=0.0, key=f"custom_filter_volume_{batch_id}")

                        if st.button("Submit Custom Filter", key=f"submit_custom_filter_{batch_id}"):
                            st.session_state.selections.append(([{
                                "unit_op": batch_id,
                                "operation": "filtration",
                                "material": "N/A",
                                "input_volume": 0,
                                "actual_volume": custom_volume,
                                "accumulated_volume": custom_volume,
                                "custom": True
                            }], custom_filter_name))
                            st.success(f"Custom filter '{custom_filter_name}' added.")

                    if uploaded_filter_file:
                        filter_df = load_filter_data(uploaded_filter_file)

                        mass = st.number_input("Mass (kg)", min_value=0.0, key=f"mass_{batch_id}")
                        bulk_density = st.number_input("Bulk density (kg/m³)", min_value=0.0, key=f"bd_{batch_id}")

                        filter_property = st.selectbox(
                            "Select a filter-specific property",
                            ["specific cake resistance (m/kg)", "rate of cake buildup", "settling rate"],
                            key=f"filter_prop_{batch_id}"
                        )

                        filter_types_required = []
                        val = 0

                        if filter_property == "specific cake resistance (m/kg)":
                            val = st.number_input("Enter specific cake resistance (m/kg)", min_value=0.0, key=f"resistance_{batch_id}")
                            if 1e7 <= val < 1e8:
                                filter_types_required = ["CENTRIFUGE", "NUTSCHE"]
                            elif 1e8 <= val < 1e10:
                                filter_types_required = ["CENTRIFUGE", "ANFD", "RPF", "VNF"]
                            elif val >= 1e10:
                                filter_types_required = ["CENTRIFUGE", "NUTSCHE"]

                        elif filter_property == "rate of cake buildup":
                            unit = st.selectbox("Select unit for rate of cake buildup", ["cm/sec", "cm/min", "cm/hr"], key=f"buildup_unit_{batch_id}")
                            val = st.number_input(f"Enter rate of cake buildup ({unit})", min_value=0.0, key=f"buildup_val_{batch_id}")
                            if unit == "cm/sec" and 0.1 <= val <= 10:
                                filter_types_required = ["CENTRIFUGE", "NUTSCHE"]
                            elif unit == "cm/min" and 0.1 <= val <= 10:
                                filter_types_required = ["CENTRIFUGE", "ANFD", "RPF"]
                            elif unit == "cm/hr" and 0.1 <= val <= 10:
                                filter_types_required = ["ANFD"]

                        elif filter_property == "settling rate":
                            val = st.number_input("Enter settling rate (cm/sec)", min_value=0.0, key=f"settling_{batch_id}")
                            if val > 5:
                                filter_types_required = ["CENTRIFUGE", "NUTSCHE"]
                            elif 0.1 <= val <= 5:
                                filter_types_required = ["ANFD", "RPF"]
                            elif val < 0.1:
                                filter_types_required = ["ANFD"]

                        if st.button(f"Submit Filtration Operation {batch_id}", key=f"submit_{batch_id}"):
                            user_input = {
                                "ph_condition": ph_condition,
                                "corrosion_rate": corrosion_rate,
                                "coupon_materials": coupon_materials,
                                "temperature": temperature,
                                "bulk_density": bulk_density,
                                "mass": mass
                            }

                            matched_df = filter_filters(filter_df.copy(), user_input, filter_types_required)

                            if not matched_df.empty:
                                # Calculate volume and cake height
                                volume_L = mass / bulk_density * 1000 if bulk_density > 0 else 0
                                volume_m = mass / bulk_density if bulk_density > 0 else 0
                                if "area" in matched_df.columns:
                                    matched_df["Cake Height (cm)"] = matched_df["area"].apply(lambda a: round((volume_m * 100 / a) if a > 0 else 0, 2))

                                st.success("Matching filters found")
                                st.dataframe(matched_df)

                                filter_id_col = next((col for col in matched_df.columns if col.strip().lower() in ["equipment id", "filter id", "id"]), None)
                                if filter_id_col:
                                    filter_options = matched_df[filter_id_col].astype(str).tolist()
                                else:
                                    st.warning("No suitable ID column found in filters. Using index.")
                                    filter_options = matched_df.index.astype(str).tolist()

                                selected_filter = st.selectbox("Select one filter to use:", filter_options, key=f"sel_filter_{batch_id}")

                                st.session_state.selections.append(([
                                    {
                                        "unit_op": batch_id,
                                        "operation": "filtration",
                                        "material": "N/A",
                                        "input_volume": 0,
                                        "actual_volume": volume_L,
                                        "accumulated_volume": volume_L
                                    }
                                ], selected_filter))
                            else:
                                st.warning("No matching filters found.")

                # ---------- DRYING OPERATION ----------
                elif unit_op_type == "drying":
                    uploaded_dryer_file = st.file_uploader(f"Upload Dryer Database (for Unit Operation {batch_id})", type=["xlsx"], key=f"upload_dryer_{batch_id}")

                    st.markdown("### Or manually enter custom dryer (not in database)")
                    custom_dryer = st.checkbox("Enter custom dryer details", key=f"custom_dryer_chk_{batch_id}")
                    if custom_dryer:
                        custom_dryer_name = st.text_input("Dryer ID or Name", key=f"custom_dryer_name_{batch_id}")
                        custom_capacity = st.number_input("Drying Volume (L)", min_value=0.0, key=f"custom_dryer_capacity_{batch_id}")

                        if st.button("Submit Custom Dryer", key=f"submit_custom_dryer_{batch_id}"):
                            st.session_state.selections.append(([{
                                "unit_op": batch_id,
                                "operation": "drying",
                                "material": "N/A",
                                "input_volume": 0,
                                "actual_volume": custom_capacity,
                                "accumulated_volume": custom_capacity,
                                "custom": True
                            }], custom_dryer_name))
                            st.success(f"Custom dryer '{custom_dryer_name}' added.")

                    if uploaded_dryer_file:
                        dryer_df = load_dryer_data(uploaded_dryer_file)

                        volume_L = st.number_input("Volume (L)", min_value=0.0, key=f"vol_dry_{batch_id}")

                        if st.button(f"Submit Drying Operation {batch_id}", key=f"submit_dry_{batch_id}"):
                            user_input = {
                                "ph_condition": ph_condition,
                                "corrosion_rate": corrosion_rate,
                                "coupon_materials": coupon_materials,
                                "temperature": temperature,
                                "volume": volume_L
                            }

                            matched_df = filter_dryers(dryer_df.copy(), user_input)
                            if "area" in matched_df.columns:
                                matched_df["Cake Height (cm)"] = matched_df["area"].apply(lambda a: round((volume_L * 0.1 / a) if a > 0 else 0, 2))

                            if not matched_df.empty:
                                st.success("Matching dryers found")
                                st.dataframe(matched_df)
                                selected_dryer = st.selectbox("Select one dryer to use:", matched_df["equipment id"].tolist() if "equipment id" in matched_df.columns else matched_df.index.astype(str), key=f"sel_dryer_{batch_id}")
                                st.session_state.selections.append(([{
                                    "unit_op": batch_id,
                                    "operation": "drying",
                                    "material": "N/A",
                                    "input_volume": 0,
                                    "actual_volume": volume_L,
                                    "accumulated_volume": volume_L
                                }], selected_dryer))
                            else:
                                st.warning("No matching dryers found.")

    # Export Excel summary
    if st.session_state.selections:
        excel_buffer = export_steps_to_excel(st.session_state.selections)
        st.download_button("Download Steps Summary", data=excel_buffer.getvalue(), file_name="unit_op_steps.xlsx")

if __name__ == "__main__":
    main()


